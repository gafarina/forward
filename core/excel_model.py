"""
Constructor del libro Excel del valorizador de forwards FX.

Este módulo genera un `.xlsx` **con fórmulas vivas**: salvo los datos de
entrada (curvas, contratos, parámetros) y las columnas de referencia de la hoja
`Reconciliación`, todo lo demás son fórmulas nativas de Excel que se recalculan
al cambiar cualquier celda azul. No hay VBA, ni macros, ni funciones definidas
por el usuario: el archivo abre en Excel, LibreOffice, Google Sheets o Numbers.

No depende de Django. Lo usan:

* `scripts/build_excel_model.py`            — generación del libro plantilla.
* `valorizador/services/excel_export.py`    — exportación de una valorización.

Réplica del motor
-----------------
Las fórmulas de la hoja `Valorización` reproducen exactamente
`core.valuation.price_contract` con `interp="Lineal"`, extrapolación
seleccionable y capitalización compuesta:

    ε          = +1 Venta, −1 Compra
    F_mercado  = interpolación lineal en la curva de outrights
    tasa       = interpolación lineal en la curva de descuento
    DF         = (1 + tasa/100) ^ (−t)          con t = días / Base_Anual
    MtM        = ε · (K − F_mercado) · N · DF
    Comp.spot  = ε · (S₀ − S_val) · N · DF
    Puntos fwd = MtM − Componente spot
    Delta      = −ε · N · DF
    DV01       = MtM(tasa + 1 bp) − MtM
    Theta 1d   = MtM(t+1, curvas congeladas) − MtM

Interpolación / extrapolación en fórmula nativa
-----------------------------------------------
El nodo inferior se localiza con ``COUNTIF`` en vez de ``MATCH``: el bloque de
curvas tiene 60 celdas reservadas y ``MATCH(...;1)`` sobre un rango con celdas
vacías al final devuelve resultados dependientes de la implementación (las
vacías se leen como cero y rompen la monotonía que MATCH exige). ``COUNTIF``
ignora las celdas vacías, de modo que

    k = MAX(1; MIN(n−1; COUNTIF(dias; "<=" & x)))

es correcto en los tres regímenes:

* x < primer nodo   → COUNTIF = 0 → k = 1     → extrapola con los dos primeros
* x dentro de rango → k = índice del nodo inferior
* x ≥ último nodo   → COUNTIF = n → k = n−1   → extrapola con los dos últimos

y la misma expresión de pendiente sirve para los tres:

    y = INDICE(val;k) + (x − INDICE(dias;k)) ·
        (INDICE(val;k+1) − INDICE(val;k)) / (INDICE(dias;k+1) − INDICE(dias;k))

La extrapolación **plana** se obtiene sin una segunda fórmula: basta acotar el
plazo consultado al rango de nodos (`MIN(MAX(x; x₁); xₙ)`). Por eso las hojas
tienen una columna auxiliar "plazo efectivo".

Todas las fórmulas se escriben en inglés (INDEX, MATCH, COUNTIF, IFERROR…),
que es como se almacenan en el formato OOXML; Excel las muestra traducidas al
idioma de la instalación.
"""

from __future__ import annotations

import io
from dataclasses import dataclass, field
from datetime import date, datetime

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.hyperlink import Hyperlink

__all__ = [
    "WorkbookData",
    "ContractRow",
    "VERSION",
    "compute_reference",
    "build_workbook_object",
    "build_workbook_bytes",
    "save_workbook",
    "demo_data",
]

VERSION = "2.0"

# ──────────────────────────────────────────────────────────────────────
# Nombres de hoja
# ──────────────────────────────────────────────────────────────────────

SH_POR = "Portada"
SH_PAR = "Parámetros"
SH_CUR = "Curvas"
SH_CON = "Contratos"
SH_VAL = "Valorización"
SH_SEN = "Sensibilidad"
SH_GRI = "Griegas"
SH_REC = "Reconciliación"
SH_MET = "Metodología"

ALL_SHEETS = (SH_POR, SH_PAR, SH_CUR, SH_CON, SH_VAL, SH_SEN, SH_GRI, SH_REC, SH_MET)


def q(sheet: str) -> str:
    """Referencia de hoja siempre entrecomillada (hay nombres con acento)."""
    return f"'{sheet}'"


# ──────────────────────────────────────────────────────────────────────
# Paleta y estilos
# ──────────────────────────────────────────────────────────────────────

C_HEADER = "4F2D7F"        # morado corporativo
C_HEADER_SOFT = "7A5AA6"
C_INPUT = "DCE9F7"         # azul claro: celda editable
C_CALC = "F5F5F5"          # gris muy claro: fórmula
C_TOTAL = "EDE7F4"
C_NEG = "C00000"
C_OK = "1E7B34"

FMT_CLP = "#,##0"
FMT_CLP2 = "#,##0.00"
FMT_PRICE = "0.0000"
FMT_RATE = "0.000000"
FMT_DF = "0.00000000"
FMT_YF = "0.00000000"
FMT_DATE = "dd-mm-yyyy"
FMT_INT = "#,##0"
FMT_PCTLBL = '0.0"%"'

FILL_HEADER = PatternFill("solid", fgColor=C_HEADER)
FILL_HEADER_SOFT = PatternFill("solid", fgColor=C_HEADER_SOFT)
FILL_INPUT = PatternFill("solid", fgColor=C_INPUT)
FILL_CALC = PatternFill("solid", fgColor=C_CALC)
FILL_TOTAL = PatternFill("solid", fgColor=C_TOTAL)

FONT_TITLE = Font(name="Calibri", size=16, bold=True, color=C_HEADER)
FONT_SUB = Font(name="Calibri", size=10, italic=True, color="595959")
FONT_HEADER = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
FONT_BOLD = Font(name="Calibri", size=10, bold=True)
FONT_BASE = Font(name="Calibri", size=10)
FONT_LINK = Font(name="Calibri", size=10, color="0563C1", underline="single")

_thin = Side(style="thin", color="BFBFBF")
BORDER_ALL = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
BORDER_TOP = Border(top=Side(style="medium", color=C_HEADER))

ALIGN_C = Alignment(horizontal="center", vertical="center")
ALIGN_L = Alignment(horizontal="left", vertical="center")
ALIGN_R = Alignment(horizontal="right", vertical="center")
ALIGN_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)


# ──────────────────────────────────────────────────────────────────────
# Datos de entrada
# ──────────────────────────────────────────────────────────────────────

@dataclass
class ContractRow:
    """Una línea de la hoja `Contratos`."""

    folio: str = ""
    counterparty: str = ""
    cartera: str = ""
    side: str = "Venta"
    modality: str = "Compensacion"
    base_ccy: str = "USD"
    notional: float = 0.0
    fwd_price: float = 0.0          # K
    spot_inicio: float = 0.0        # S₀
    start_date: date | None = None
    maturity_date: date | None = None

    @property
    def sign(self) -> int:
        return 1 if str(self.side).strip().lower().startswith("v") else -1


@dataclass
class WorkbookData:
    """Todo lo que el libro necesita para construirse."""

    valuation_date: date
    spot: float
    fwd_nodes: list[tuple[float, float]] = field(default_factory=list)
    desc_nodes: list[tuple[float, float]] = field(default_factory=list)
    contracts: list[ContractRow] = field(default_factory=list)
    base_anual: int = 360
    extrap: str = "Lineal"          # "Lineal" | "Plana"
    day_count: str = "ACT/360"
    compounding: str = "Compuesta"
    label: str = "Cartera"
    titulo: str = "Valorizador de Forwards FX USD/CLP"
    fuente: str = ""
    shock_max: float = 5.0
    referencia: list[dict] | None = None   # MtM del motor Python por contrato
    advertencias: list[str] = field(default_factory=list)

    # -- derivados ----------------------------------------------------
    @property
    def n_rows(self) -> int:
        """Filas preparadas en Contratos / Valorización (mínimo 100)."""
        return max(100, len(self.contracts))

    @property
    def n_nodes(self) -> int:
        """Celdas reservadas por bloque de curva (mínimo 60)."""
        return max(60, len(self.fwd_nodes), len(self.desc_nodes))


# ──────────────────────────────────────────────────────────────────────
# Referencia del motor Python (para la hoja Reconciliación)
# ──────────────────────────────────────────────────────────────────────

def compute_reference(data: WorkbookData) -> list[dict]:
    """
    Valoriza los contratos con el motor Python (`core.curves`) sin redondear.

    Se usa la misma clase `Curve` que alimenta la aplicación web, de modo que
    la reconciliación compara de verdad Excel contra motor y no Excel contra
    una segunda implementación escrita a mano.
    """
    from .curves import Curve

    if not data.fwd_nodes or not data.desc_nodes:
        return []

    fwd = Curve(
        "FWD",
        [float(x) for x, _ in data.fwd_nodes],
        [float(y) for _, y in data.fwd_nodes],
        interp="Lineal",
        extrap=data.extrap,
    )
    dsc = Curve(
        "DESC",
        [float(x) for x, _ in data.desc_nodes],
        [float(y) for _, y in data.desc_nodes],
        interp="Lineal",
        extrap=data.extrap,
    )

    base = float(data.base_anual)
    out: list[dict] = []
    for c in data.contracts:
        if not c.maturity_date or not c.notional:
            out.append({})
            continue
        d = (c.maturity_date - data.valuation_date).days
        t = d / base
        F = fwd.value(d)
        r = dsc.value(d)
        df = (1.0 + r / 100.0) ** (-t)
        eps = c.sign
        K = float(c.fwd_price)
        N = float(c.notional)
        S0 = float(c.spot_inicio or 0.0)
        mtm = eps * (K - F) * N * df
        spot_c = eps * (S0 - float(data.spot)) * N * df if S0 > 0 else 0.0
        df_up = (1.0 + (r + 0.01) / 100.0) ** (-t)
        out.append(
            {
                "folio": c.folio,
                "counterparty": c.counterparty,
                "days": d,
                "year_fraction": t,
                "fwd_mkt": F,
                "disc_rate": r,
                "disc_factor": df,
                "mtm": mtm,
                "spot_component": spot_c,
                "fwd_points": mtm - spot_c,
                "delta": -eps * N * df,
                "dv01": eps * (K - F) * N * df_up - mtm,
            }
        )
    return out


# ──────────────────────────────────────────────────────────────────────
# Utilidades de escritura
# ──────────────────────────────────────────────────────────────────────

def _title(ws, text: str, subtitle: str = "", width: int = 10) -> None:
    ws["A1"] = text
    ws["A1"].font = FONT_TITLE
    ws.row_dimensions[1].height = 24
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = FONT_SUB


def _hdr(ws, row: int, col: int, text: str, width: float | None = None) -> None:
    c = ws.cell(row=row, column=col, value=text)
    c.fill = FILL_HEADER
    c.font = FONT_HEADER
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = BORDER_ALL
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width


def _label(ws, ref: str, text: str, bold: bool = True) -> None:
    ws[ref] = text
    ws[ref].font = FONT_BOLD if bold else FONT_BASE


def _input(ws, ref: str, value, fmt: str | None = None) -> None:
    c = ws[ref]
    c.value = value
    c.fill = FILL_INPUT
    c.border = BORDER_ALL
    c.font = FONT_BASE
    c.alignment = ALIGN_C
    if fmt:
        c.number_format = fmt


def _calc(ws, ref: str, value, fmt: str | None = None, bold: bool = False) -> None:
    c = ws[ref]
    c.value = value
    c.fill = FILL_CALC
    c.border = BORDER_ALL
    c.font = FONT_BOLD if bold else FONT_BASE
    c.alignment = ALIGN_C
    if fmt:
        c.number_format = fmt


def _widths(ws, spec: dict[str, float]) -> None:
    for col, w in spec.items():
        ws.column_dimensions[col].width = w


def _dv(ws, opciones: str, sqref: str, titulo: str = "Valor no permitido") -> None:
    dv = DataValidation(
        type="list", formula1=f'"{opciones}"', allow_blank=True, showDropDown=False
    )
    dv.error = f"Elija uno de: {opciones}"
    dv.errorTitle = titulo
    ws.add_data_validation(dv)
    dv.add(sqref)


# ──────────────────────────────────────────────────────────────────────
# Hoja: Portada
# ──────────────────────────────────────────────────────────────────────

def _sheet_portada(wb: Workbook, data: WorkbookData, hojas: tuple[str, ...]) -> None:
    ws = wb[SH_POR]
    ws.sheet_view.showGridLines = False
    _widths(ws, {"A": 3, "B": 30, "C": 78, "D": 18})

    ws["B2"] = data.titulo
    ws["B2"].font = Font(name="Calibri", size=22, bold=True, color=C_HEADER)
    ws.row_dimensions[2].height = 30
    ws["B3"] = "Valorización a mercado de forwards de moneda con fórmulas vivas de Excel"
    ws["B3"].font = FONT_SUB

    fila = 5
    meta = [
        ("Versión del modelo", VERSION),
        ("Generado el", datetime.now().strftime("%d-%m-%Y %H:%M")),
        ("Fecha de valorización", data.valuation_date.strftime("%d-%m-%Y")),
        ("Cartera / etiqueta", data.label),
        ("Spot de valorización (CLP/USD)", f"{data.spot:,.4f}".replace(",", "@").replace(".", ",").replace("@", ".")),
        ("Convención de días", data.day_count),
        ("Capitalización", data.compounding),
        ("Extrapolación de curvas", data.extrap),
        ("Contratos cargados", str(len([c for c in data.contracts if c.notional]))),
        ("Fuente de datos", data.fuente or "Carga manual"),
    ]
    for k, v in meta:
        ws.cell(row=fila, column=2, value=k).font = FONT_BOLD
        c = ws.cell(row=fila, column=3, value=v)
        c.font = FONT_BASE
        c.alignment = ALIGN_L
        fila += 1

    fila += 1
    ws.cell(row=fila, column=2, value="QUÉ HACE ESTE LIBRO").font = Font(
        name="Calibri", size=12, bold=True, color=C_HEADER
    )
    fila += 1
    desc = (
        "Replica celda por celda el motor de valorización del sistema. Toma una curva de forwards "
        "outright y una curva de descuento, interpola linealmente en ambas (con extrapolación lineal "
        "fuera del rango de nodos), descuenta y calcula el MtM de cada contrato, su descomposición en "
        "componente spot y puntos forward, y las sensibilidades. Todo son fórmulas de Excel: si usted "
        "cambia el spot, un nodo de curva o el nocional de un contrato, el libro entero se recalcula. "
        "No contiene macros ni funciones definidas por el usuario."
    )
    ws.cell(row=fila, column=2, value=desc)
    ws.merge_cells(start_row=fila, start_column=2, end_row=fila + 3, end_column=4)
    ws.cell(row=fila, column=2).alignment = ALIGN_WRAP
    ws.cell(row=fila, column=2).font = FONT_BASE
    fila += 5

    ws.cell(row=fila, column=2, value="CONVENCIÓN DE COLORES").font = Font(
        name="Calibri", size=12, bold=True, color=C_HEADER
    )
    fila += 1
    leyenda = [
        (FILL_INPUT, "Celda de entrada", "Azul claro con borde. Es editable: cámbiela con confianza."),
        (FILL_CALC, "Celda calculada", "Gris muy claro. Contiene una fórmula; no la sobrescriba."),
        (FILL_HEADER, "Encabezado", "Morado con texto blanco. Rótulos de tabla."),
        (FILL_TOTAL, "Totales", "Morado pálido. Sumas y agregados de cartera."),
    ]
    for fill, nombre, expl in leyenda:
        c = ws.cell(row=fila, column=2, value=nombre)
        c.fill = fill
        c.border = BORDER_ALL
        c.font = FONT_HEADER if fill is FILL_HEADER else FONT_BOLD
        c.alignment = ALIGN_C
        e = ws.cell(row=fila, column=3, value=expl)
        e.font = FONT_BASE
        e.alignment = ALIGN_L
        fila += 1

    fila += 1
    ws.cell(row=fila, column=2, value="ÍNDICE").font = Font(
        name="Calibri", size=12, bold=True, color=C_HEADER
    )
    fila += 1
    indice = {
        SH_PAR: "Fecha de valorización, spot, base anual y método de extrapolación.",
        SH_CUR: "Nodos de la curva forward (outright) y de la curva de descuento.",
        SH_CON: "Ficha de cada operación: nocional, precio pactado, spot inicial y fechas.",
        SH_VAL: "Núcleo del modelo. Una fila por contrato, íntegramente en fórmulas.",
        SH_SEN: "Matriz 5×5 de MtM ante desplazamientos de spot y de curva forward.",
        SH_GRI: "Delta, DV01 y theta agregados, con desglose por contraparte.",
        SH_REC: "Excel contra motor Python, contrato por contrato.",
        SH_MET: "Fórmulas, convenciones y supuestos en notación legible.",
    }
    for hoja, expl in indice.items():
        if hoja not in hojas:
            continue
        c = ws.cell(row=fila, column=2, value=hoja)
        c.font = FONT_LINK
        c.alignment = ALIGN_L
        c.hyperlink = Hyperlink(ref=c.coordinate, location=f"{q(hoja)}!A1", display=hoja)
        e = ws.cell(row=fila, column=3, value=expl)
        e.font = FONT_BASE
        e.alignment = ALIGN_L
        fila += 1

    fila += 1
    nota = (
        "Advertencia: al abrir el archivo, algunas aplicaciones no recalculan las fórmulas hasta que "
        "se fuerza el recálculo. En Excel, pulse Ctrl+Alt+F9 si ve celdas vacías donde espera números."
    )
    ws.cell(row=fila, column=2, value=nota).font = FONT_SUB
    ws.merge_cells(start_row=fila, start_column=2, end_row=fila, end_column=4)

    if data.advertencias:
        fila += 2
        c = ws.cell(row=fila, column=2, value="ADVERTENCIAS DE ESTA EXPORTACIÓN")
        c.font = Font(name="Calibri", size=12, bold=True, color="9C5700")
        fila += 1
        for texto in data.advertencias:
            c = ws.cell(row=fila, column=2, value="· " + texto)
            c.font = Font(name="Calibri", size=10, color="9C5700")
            c.alignment = ALIGN_WRAP
            c.fill = PatternFill("solid", fgColor="FFF2CC")
            ws.merge_cells(start_row=fila, start_column=2, end_row=fila, end_column=4)
            ws.row_dimensions[fila].height = max(15, 14 * (len(texto) // 100 + 1))
            fila += 1


# ──────────────────────────────────────────────────────────────────────
# Hoja: Parámetros
# ──────────────────────────────────────────────────────────────────────

def _sheet_parametros(wb: Workbook, data: WorkbookData) -> None:
    ws = wb[SH_PAR]
    ws.sheet_view.showGridLines = False
    _widths(ws, {"A": 3, "B": 34, "C": 20, "D": 4, "E": 74})
    _title(ws, "Parámetros de valorización",
           "Celdas azules: entradas del usuario. Todo el libro depende de ellas.")

    _hdr(ws, 4, 2, "Parámetro")
    _hdr(ws, 4, 3, "Valor")
    _hdr(ws, 4, 5, "Comentario")

    filas = [
        ("Fecha de valorización", "B5", "C5", data.valuation_date, FMT_DATE,
         "Fecha a la que se descuenta. Los días al vencimiento son días corridos "
         "entre esta fecha y el vencimiento de cada contrato."),
        ("Spot de valorización (CLP/USD)", "B6", "C6", float(data.spot), FMT_PRICE,
         "Tipo de cambio observado hoy. Entra en el componente spot y en la matriz de sensibilidad."),
        ("Base anual (días)", "B7", "C7", int(data.base_anual), "0",
         "Denominador de la fracción de año. 360 para ACT/360 (convención local), 365 para ACT/365."),
        ("Método de extrapolación", "B8", "C8", data.extrap, None,
         "Lineal: prolonga la pendiente de los dos nodos extremos (reproduce la planilla original). "
         "Plana: mantiene constante el valor del nodo extremo."),
        ("Convención de días", "B9", "C9", data.day_count, None,
         "Informativo. La fracción de año efectiva es días corridos / Base anual."),
        ("Capitalización", "B10", "C10", data.compounding, None,
         "Compuesta: DF = (1 + tasa/100)^(−t). Es la convención del mercado CLP."),
        ("Etiqueta de la cartera", "B11", "C11", data.label, None,
         "Sólo descriptivo; aparece en la portada."),
    ]
    for etiqueta, ref_lbl, ref_val, valor, fmt, coment in filas:
        _label(ws, ref_lbl, etiqueta)
        ws[ref_lbl].border = BORDER_ALL
        _input(ws, ref_val, valor, fmt)
        cc = ws[f"E{ref_val[1:]}"]
        cc.value = coment
        cc.font = FONT_SUB
        cc.alignment = ALIGN_WRAP
        ws.row_dimensions[int(ref_val[1:])].height = 30

    _dv(ws, "360,365", "C7", "Base anual")
    _dv(ws, "Lineal,Plana", "C8", "Extrapolación")
    _dv(ws, "ACT/360,ACT/365,30/360,30E/360,ACT/ACT", "C9", "Convención de días")
    _dv(ws, "Compuesta,Simple,Continua", "C10", "Capitalización")

    ws["B13"] = "Nombres definidos disponibles en todo el libro"
    ws["B13"].font = Font(name="Calibri", size=11, bold=True, color=C_HEADER)
    nombres = [
        ("Fecha_Valorizacion", "Parámetros!C5"),
        ("Spot_Valorizacion", "Parámetros!C6"),
        ("Base_Anual", "Parámetros!C7"),
        ("Metodo_Extrapolacion", "Parámetros!C8"),
        ("Curva_Fwd_Dias / Curva_Fwd_Val", "Curvas: nodos de la curva de outrights"),
        ("Curva_Desc_Dias / Curva_Desc_Val", "Curvas: nodos de la curva de descuento"),
        ("N_Fwd / N_Desc", "Curvas: cantidad de nodos activos de cada curva"),
        ("Shock_Max", "Sensibilidad: amplitud máxima de los escenarios (%)"),
    ]
    fila = 14
    for n, r in nombres:
        ws.cell(row=fila, column=2, value=n).font = FONT_BASE
        ws.cell(row=fila, column=3, value="").font = FONT_BASE
        ws.cell(row=fila, column=5, value=r).font = FONT_SUB
        fila += 1

    aviso = (
        "El motor Python usa por defecto extrapolación plana; la planilla Cordada original usa "
        "extrapolación LINEAL, y por eso ese es el valor por defecto acá. Cambiarlo altera todos los "
        "contratos cuyo plazo caiga fuera del rango de nodos de alguna de las dos curvas."
    )
    ws.cell(row=fila + 1, column=2, value=aviso).font = FONT_SUB
    ws.merge_cells(start_row=fila + 1, start_column=2, end_row=fila + 2, end_column=5)
    ws.cell(row=fila + 1, column=2).alignment = ALIGN_WRAP

    wb.defined_names.add(DefinedName("Fecha_Valorizacion", attr_text=f"{q(SH_PAR)}!$C$5"))
    wb.defined_names.add(DefinedName("Spot_Valorizacion", attr_text=f"{q(SH_PAR)}!$C$6"))
    wb.defined_names.add(DefinedName("Base_Anual", attr_text=f"{q(SH_PAR)}!$C$7"))
    wb.defined_names.add(DefinedName("Metodo_Extrapolacion", attr_text=f"{q(SH_PAR)}!$C$8"))


# ──────────────────────────────────────────────────────────────────────
# Hoja: Curvas
# ──────────────────────────────────────────────────────────────────────

CUR_FIRST_ROW = 7


def _sheet_curvas(wb: Workbook, data: WorkbookData) -> None:
    ws = wb[SH_CUR]
    ws.sheet_view.showGridLines = False
    n = data.n_nodes
    last = CUR_FIRST_ROW + n - 1

    _widths(ws, {"A": 12, "B": 16, "C": 22, "D": 4, "E": 12, "F": 16, "G": 4, "H": 62})
    _title(ws, "Curvas de mercado",
           "Cargue los nodos desde la primera fila del bloque, en orden ascendente de plazo y sin dejar huecos.")

    _label(ws, "A3", "Nodos forward:")
    _calc(ws, "B3", "=COUNT(Curva_Fwd_Dias)", "0", bold=True)
    _label(ws, "E3", "Nodos descuento:")
    _calc(ws, "F3", "=COUNT(Curva_Desc_Dias)", "0", bold=True)

    ws["H3"] = ("N_Fwd y N_Desc cuentan las celdas de plazo con número. Son los que usan las fórmulas "
                "de interpolación para ubicar los nodos extremos, de modo que agregar o quitar nodos "
                "no exige tocar ninguna fórmula.")
    ws["H3"].font = FONT_SUB
    ws["H3"].alignment = ALIGN_WRAP
    ws.merge_cells("H3:H6")

    ws.merge_cells(start_row=5, start_column=1, end_row=5, end_column=3)
    c = ws.cell(row=5, column=1, value="CURVA FORWARD USD/CLP (outright)")
    c.fill = FILL_HEADER_SOFT
    c.font = FONT_HEADER
    c.alignment = ALIGN_C
    ws.merge_cells(start_row=5, start_column=5, end_row=5, end_column=6)
    c = ws.cell(row=5, column=5, value="CURVA DE DESCUENTO CLP")
    c.fill = FILL_HEADER_SOFT
    c.font = FONT_HEADER
    c.alignment = ALIGN_C

    _hdr(ws, 6, 1, "Días")
    _hdr(ws, 6, 2, "Outright")
    _hdr(ws, 6, 3, "Puntos fwd implícitos")
    _hdr(ws, 6, 5, "Días")
    _hdr(ws, 6, 6, "Tasa (%)")

    fwd = list(data.fwd_nodes)
    dsc = list(data.desc_nodes)

    for i in range(n):
        r = CUR_FIRST_ROW + i
        xa = ws.cell(row=r, column=1)
        ya = ws.cell(row=r, column=2)
        if i < len(fwd):
            xa.value = float(fwd[i][0])
            ya.value = float(fwd[i][1])
        for cc, fmt in ((xa, FMT_INT), (ya, FMT_PRICE)):
            cc.fill = FILL_INPUT
            cc.border = BORDER_ALL
            cc.font = FONT_BASE
            cc.alignment = ALIGN_C
            cc.number_format = fmt
        _calc(ws, f"C{r}", f'=IF(A{r}="","",B{r}-Spot_Valorizacion)', FMT_PRICE)

        xb = ws.cell(row=r, column=5)
        yb = ws.cell(row=r, column=6)
        if i < len(dsc):
            xb.value = float(dsc[i][0])
            yb.value = float(dsc[i][1])
        for cc, fmt in ((xb, FMT_INT), (yb, FMT_RATE)):
            cc.fill = FILL_INPUT
            cc.border = BORDER_ALL
            cc.font = FONT_BASE
            cc.alignment = ALIGN_C
            cc.number_format = fmt

    ws.freeze_panes = "A7"

    wb.defined_names.add(DefinedName("Curva_Fwd_Dias", attr_text=f"{q(SH_CUR)}!$A${CUR_FIRST_ROW}:$A${last}"))
    wb.defined_names.add(DefinedName("Curva_Fwd_Val", attr_text=f"{q(SH_CUR)}!$B${CUR_FIRST_ROW}:$B${last}"))
    wb.defined_names.add(DefinedName("Curva_Desc_Dias", attr_text=f"{q(SH_CUR)}!$E${CUR_FIRST_ROW}:$E${last}"))
    wb.defined_names.add(DefinedName("Curva_Desc_Val", attr_text=f"{q(SH_CUR)}!$F${CUR_FIRST_ROW}:$F${last}"))
    wb.defined_names.add(DefinedName("N_Fwd", attr_text=f"{q(SH_CUR)}!$B$3"))
    wb.defined_names.add(DefinedName("N_Desc", attr_text=f"{q(SH_CUR)}!$F$3"))


# ──────────────────────────────────────────────────────────────────────
# Hoja: Contratos
# ──────────────────────────────────────────────────────────────────────

CON_FIRST_ROW = 4
CON_COLS = [
    ("Folio", "A", 12),
    ("Contraparte", "B", 20),
    ("Cartera", "C", 16),
    ("Operación", "D", 12),
    ("Modalidad", "E", 16),
    ("Moneda", "F", 10),
    ("Nocional", "G", 16),
    ("Precio Fwd Pactado", "H", 16),
    ("Spot al Inicio", "I", 14),
    ("Fecha Inicio", "J", 14),
    ("Fecha Vencimiento", "K", 16),
]


def _sheet_contratos(wb: Workbook, data: WorkbookData) -> None:
    ws = wb[SH_CON]
    ws.sheet_view.showGridLines = False
    nrows = data.n_rows
    last = CON_FIRST_ROW + nrows - 1

    _title(ws, "Contratos forward",
           "Una fila por operación. Todas las celdas son de entrada; la hoja Valorización las lee posición a posición.")

    for nombre, col, w in CON_COLS:
        idx = ws[f"{col}1"].column
        _hdr(ws, 3, idx, nombre, w)
    ws.row_dimensions[3].height = 28

    fmts = {
        "G": FMT_CLP, "H": FMT_PRICE, "I": FMT_PRICE,
        "J": FMT_DATE, "K": FMT_DATE,
    }
    for i in range(nrows):
        r = CON_FIRST_ROW + i
        c = data.contracts[i] if i < len(data.contracts) else None
        vals = {
            "A": c.folio if c else None,
            "B": c.counterparty if c else None,
            "C": c.cartera if c else None,
            "D": c.side if c else None,
            "E": c.modality if c else None,
            "F": c.base_ccy if c else None,
            "G": float(c.notional) if c and c.notional else None,
            "H": float(c.fwd_price) if c and c.fwd_price else None,
            "I": float(c.spot_inicio) if c and c.spot_inicio else None,
            "J": c.start_date if c else None,
            "K": c.maturity_date if c else None,
        }
        for col in "ABCDEFGHIJK":
            cell = ws[f"{col}{r}"]
            cell.value = vals[col]
            cell.fill = FILL_INPUT
            cell.border = BORDER_ALL
            cell.font = FONT_BASE
            cell.alignment = ALIGN_C if col in "DEF" else (
                ALIGN_R if col in "GHI" else ALIGN_L
            )
            if col in fmts:
                cell.number_format = fmts[col]

    _dv(ws, "Compra,Venta", f"D{CON_FIRST_ROW}:D{last}", "Operación")
    _dv(ws, "Compensacion,Entrega", f"E{CON_FIRST_ROW}:E{last}", "Modalidad")
    _dv(ws, "USD,EUR,CLP,UF", f"F{CON_FIRST_ROW}:F{last}", "Moneda")

    ws.auto_filter.ref = f"A3:K{last}"
    ws.freeze_panes = "C4"


# ──────────────────────────────────────────────────────────────────────
# Hoja: Valorización
# ──────────────────────────────────────────────────────────────────────

VAL_FIRST_ROW = 4

VAL_COLS = [
    ("A", "Folio", 12, None),
    ("B", "Contraparte", 20, None),
    ("C", "Operación", 11, None),
    ("D", "ε (signo)", 9, "0"),
    ("E", "Vencimiento", 13, FMT_DATE),
    ("F", "Días a Vcto", 11, "0"),
    ("G", "Fracción de año", 13, FMT_YF),
    ("H", "F mercado", 14, FMT_PRICE),
    ("I", "Tasa descuento (%)", 14, FMT_RATE),
    ("J", "Factor de descuento", 15, FMT_DF),
    ("K", "Nocional", 15, FMT_CLP),
    ("L", "K pactado", 13, FMT_PRICE),
    ("M", "S₀ inicio", 13, FMT_PRICE),
    ("N", "S valorización", 13, FMT_PRICE),
    ("O", "MtM (CLP)", 17, FMT_CLP),
    ("P", "Componente spot", 17, FMT_CLP),
    ("Q", "Puntos forward", 17, FMT_CLP),
    ("R", "Delta (CLP/CLP)", 15, FMT_CLP),
    ("S", "DV01 (CLP)", 14, FMT_CLP2),
    ("T", "Alertas", 46, None),
]

VAL_AUX = [
    ("V", "Plazo efectivo fwd", 15, "0.####"),
    ("W", "k fwd", 8, "0"),
    ("X", "Plazo efectivo desc", 16, "0.####"),
    ("Y", "k desc", 8, "0"),
    ("Z", "Coef ε·N·DF", 16, FMT_CLP),
    ("AA", "K (num)", 12, FMT_PRICE),
    ("AB", "F (num)", 12, FMT_PRICE),
    ("AC", "Plazo ef. fwd t+1", 15, "0.####"),
    ("AD", "k fwd t+1", 9, "0"),
    ("AE", "F mercado t+1", 14, FMT_PRICE),
    ("AF", "Plazo ef. desc t+1", 16, "0.####"),
    ("AG", "k desc t+1", 9, "0"),
    ("AH", "Tasa t+1 (%)", 13, FMT_RATE),
    ("AI", "DF t+1", 14, FMT_DF),
    ("AJ", "Theta 1 día (CLP)", 15, FMT_CLP2),
]


def _plazo_efectivo(dias_expr: str, dias_name: str, n_name: str) -> str:
    """Plazo consultado, acotado al rango de nodos si la extrapolación es plana."""
    return (
        f'IF(Metodo_Extrapolacion="Plana",'
        f"MIN(MAX({dias_expr},INDEX({dias_name},1)),INDEX({dias_name},{n_name})),"
        f"{dias_expr})"
    )


def _indice_k(x_ref: str, dias_name: str, n_name: str) -> str:
    """Índice del nodo inferior, válido dentro y fuera del rango de la curva."""
    return f'MAX(1,MIN({n_name}-1,COUNTIF({dias_name},"<="&{x_ref})))'


def _interp(x_ref: str, k_ref: str, dias_name: str, val_name: str) -> str:
    """Interpolación / extrapolación lineal con pendiente entre los nodos k y k+1."""
    return (
        f"INDEX({val_name},{k_ref})"
        f"+({x_ref}-INDEX({dias_name},{k_ref}))"
        f"*(INDEX({val_name},{k_ref}+1)-INDEX({val_name},{k_ref}))"
        f"/(INDEX({dias_name},{k_ref}+1)-INDEX({dias_name},{k_ref}))"
    )


def _sheet_valorizacion(wb: Workbook, data: WorkbookData) -> None:
    ws = wb[SH_VAL]
    ws.sheet_view.showGridLines = False
    nrows = data.n_rows
    first, last = VAL_FIRST_ROW, VAL_FIRST_ROW + nrows - 1
    tot = last + 2

    _title(ws, "Valorización",
           "Hoja núcleo. Todas las celdas son fórmulas: no escriba números acá, edite Parámetros, Curvas o Contratos.")

    for col, nombre, w, _fmt in VAL_COLS:
        _hdr(ws, 3, ws[f"{col}1"].column, nombre, w)
    ws.row_dimensions[3].height = 32

    ws.cell(row=2, column=ws["V1"].column, value="AUXILIARES DE INTERPOLACIÓN Y GRIEGAS — no editar").font = Font(
        name="Calibri", size=9, italic=True, color="808080"
    )
    for col, nombre, w, _fmt in VAL_AUX:
        _hdr(ws, 3, ws[f"{col}1"].column, nombre, w)
        ws.cell(row=3, column=ws[f"{col}1"].column).fill = FILL_HEADER_SOFT

    CON = q(SH_CON)
    fmt_by_col = {c[0]: c[3] for c in VAL_COLS}
    fmt_by_col.update({c[0]: c[3] for c in VAL_AUX})

    for i in range(nrows):
        r = first + i
        guard = f'IF({CON}!$G{r}="","",'          # fila sin nocional = fila vacía

        f = {}
        f["A"] = f"{guard}{CON}!$A{r})"
        f["B"] = f"{guard}{CON}!$B{r})"
        f["C"] = f"{guard}{CON}!$D{r})"
        f["D"] = f'{guard}IF(LEFT(UPPER({CON}!$D{r}&""),1)="C",-1,1))'
        f["E"] = f"{guard}{CON}!$K{r})"
        f["F"] = f"{guard}{CON}!$K{r}-Fecha_Valorizacion)"
        f["G"] = f'{guard}$F{r}/Base_Anual)'

        # auxiliares de curva (t)
        f["V"] = f'IF($F{r}="","",{_plazo_efectivo(f"$F{r}", "Curva_Fwd_Dias", "N_Fwd")})'
        f["W"] = f'IF($V{r}="","",{_indice_k(f"$V{r}", "Curva_Fwd_Dias", "N_Fwd")})'
        f["X"] = f'IF($F{r}="","",{_plazo_efectivo(f"$F{r}", "Curva_Desc_Dias", "N_Desc")})'
        f["Y"] = f'IF($X{r}="","",{_indice_k(f"$X{r}", "Curva_Desc_Dias", "N_Desc")})'

        f["H"] = f'IF($W{r}="","",{_interp(f"$V{r}", f"$W{r}", "Curva_Fwd_Dias", "Curva_Fwd_Val")})'
        f["I"] = f'IF($Y{r}="","",{_interp(f"$X{r}", f"$Y{r}", "Curva_Desc_Dias", "Curva_Desc_Val")})'
        f["J"] = f'IF($I{r}="","",(1+$I{r}/100)^(-$G{r}))'

        f["K"] = f"{guard}{CON}!$G{r})"
        f["L"] = f"{guard}{CON}!$H{r})"
        f["M"] = f"{guard}N({CON}!$I{r}))"
        f["N"] = f"{guard}Spot_Valorizacion)"

        f["O"] = f'IF($J{r}="","",$D{r}*($L{r}-$H{r})*$K{r}*$J{r})'
        f["P"] = f'IF($J{r}="","",IF($M{r}>0,$D{r}*($M{r}-$N{r})*$K{r}*$J{r},0))'
        f["Q"] = f'IF($O{r}="","",$O{r}-$P{r})'
        f["R"] = f'IF($J{r}="","",-$D{r}*$K{r}*$J{r})'
        f["S"] = f'IF($J{r}="","",$D{r}*($L{r}-$H{r})*$K{r}*(1+($I{r}+0.01)/100)^(-$G{r})-$O{r})'

        f["T"] = (
            f"{guard}"
            f'IF($F{r}<0,"Contrato vencido a la fecha de valorización; ","")'
            f'&IF($F{r}=0,"Vence hoy; ","")'
            f'&IF(N({CON}!$I{r})<=0,"Sin spot al inicio: la descomposición no es confiable; ","")'
            f'&IF(N({CON}!$G{r})<=0,"Nocional no positivo; ","")'
            f'&IF(N({CON}!$H{r})<=0,"Precio pactado no positivo; ","")'
            f'&IF(OR($F{r}<INDEX(Curva_Fwd_Dias,1),$F{r}>INDEX(Curva_Fwd_Dias,N_Fwd)),'
            f'"Forward "&LOWER(Metodo_Extrapolacion)&"mente extrapolado; ","")'
            f'&IF(OR($F{r}<INDEX(Curva_Desc_Dias,1),$F{r}>INDEX(Curva_Desc_Dias,N_Desc)),'
            f'"Descuento extrapolado; ",""))'
        )

        # auxiliares numéricos para la matriz de sensibilidad (0 si la fila está vacía)
        f["Z"] = f'IF($J{r}="",0,$D{r}*$K{r}*$J{r})'
        f["AA"] = f'IF($J{r}="",0,$L{r})'
        f["AB"] = f'IF($J{r}="",0,$H{r})'

        # theta: un día de paso del tiempo con curvas congeladas
        f["AC"] = f'IF($F{r}="","",{_plazo_efectivo(f"($F{r}-1)", "Curva_Fwd_Dias", "N_Fwd")})'
        f["AD"] = f'IF($AC{r}="","",{_indice_k(f"$AC{r}", "Curva_Fwd_Dias", "N_Fwd")})'
        f["AE"] = f'IF($AD{r}="","",{_interp(f"$AC{r}", f"$AD{r}", "Curva_Fwd_Dias", "Curva_Fwd_Val")})'
        f["AF"] = f'IF($F{r}="","",{_plazo_efectivo(f"($F{r}-1)", "Curva_Desc_Dias", "N_Desc")})'
        f["AG"] = f'IF($AF{r}="","",{_indice_k(f"$AF{r}", "Curva_Desc_Dias", "N_Desc")})'
        f["AH"] = f'IF($AG{r}="","",{_interp(f"$AF{r}", f"$AG{r}", "Curva_Desc_Dias", "Curva_Desc_Val")})'
        f["AI"] = f'IF($AH{r}="","",(1+$AH{r}/100)^(-($F{r}-1)/Base_Anual))'
        f["AJ"] = (
            f'IF($O{r}="","",IF($F{r}<=0,0,'
            f'$D{r}*($L{r}-$AE{r})*$K{r}*$AI{r}-$O{r}))'
        )

        for col, formula in f.items():
            cell = ws[f"{col}{r}"]
            cell.value = "=" + formula
            cell.fill = FILL_CALC
            cell.border = BORDER_ALL
            cell.font = FONT_BASE
            fmt = fmt_by_col.get(col)
            if fmt:
                cell.number_format = fmt
            cell.alignment = ALIGN_L if col in ("A", "B", "T") else ALIGN_R

    # ── totales ──────────────────────────────────────────────────────
    tc = ws.cell(row=tot, column=1, value="TOTAL CARTERA")
    tc.font = Font(name="Calibri", size=10, bold=True, color=C_HEADER)
    for col in "ABCDEFGHIJKLMNOPQRST":
        c = ws[f"{col}{tot}"]
        c.fill = FILL_TOTAL
        c.border = BORDER_TOP
        if c.value is None:
            c.value = None
    for col in ("K", "O", "P", "Q", "R", "S"):
        c = ws[f"{col}{tot}"]
        c.value = f"=SUM({col}{first}:{col}{last})"
        c.number_format = FMT_CLP if col != "S" else FMT_CLP2
        c.font = Font(name="Calibri", size=10, bold=True)
        c.alignment = ALIGN_R
    c = ws[f"D{tot}"]
    c.value = f"=COUNT(D{first}:D{last})"
    c.number_format = "0"
    c.font = FONT_BOLD
    c.alignment = ALIGN_C
    ws[f"C{tot}"].value = "Contratos:"
    ws[f"C{tot}"].font = FONT_BOLD
    ws[f"C{tot}"].alignment = ALIGN_R

    ws.cell(row=tot, column=ws["AJ1"].column, value=f"=SUM(AJ{first}:AJ{last})").number_format = FMT_CLP2
    ws.cell(row=tot, column=ws["AJ1"].column).font = FONT_BOLD
    ws.cell(row=tot, column=ws["AJ1"].column).fill = FILL_TOTAL

    # ── formato condicional: MtM negativo en rojo ────────────────────
    rojo = Font(color=C_NEG, bold=False)
    for col in ("O", "P", "Q", "R", "S"):
        ws.conditional_formatting.add(
            f"{col}{first}:{col}{last}",
            CellIsRule(operator="lessThan", formula=["0"], font=rojo),
        )
    ws.conditional_formatting.add(
        f"T{first}:T{last}",
        CellIsRule(operator="notEqual", formula=['""'],
                   font=Font(color="9C5700"), fill=PatternFill("solid", fgColor="FFF2CC")),
    )

    ws.auto_filter.ref = f"A3:T{last}"
    ws.freeze_panes = "C4"
    ws.column_dimensions["U"].width = 3


# ──────────────────────────────────────────────────────────────────────
# Hoja: Sensibilidad
# ──────────────────────────────────────────────────────────────────────

def _sheet_sensibilidad(wb: Workbook, data: WorkbookData) -> None:
    ws = wb[SH_SEN]
    ws.sheet_view.showGridLines = False
    nrows = data.n_rows
    first, last = VAL_FIRST_ROW, VAL_FIRST_ROW + nrows - 1
    VAL = q(SH_VAL)
    rZ = f"{VAL}!$Z${first}:$Z${last}"
    rAA = f"{VAL}!$AA${first}:$AA${last}"
    rAB = f"{VAL}!$AB${first}:$AB${last}"

    _widths(ws, {"A": 26, "B": 20, "C": 20, "D": 20, "E": 20, "F": 20, "G": 4, "H": 60})
    _title(ws, "Matriz de sensibilidad",
           "MtM total de la cartera ante desplazamientos simultáneos del spot y de la curva forward.")

    _label(ws, "A4", "Shock máximo (%)")
    _input(ws, "B4", float(data.shock_max), "0.0")
    wb.defined_names.add(DefinedName("Shock_Max", attr_text=f"{q(SH_SEN)}!$B$4"))

    ws["H4"] = (
        "Los cinco escenarios de cada eje son −Shock_Max, −Shock_Max/2, 0, +Shock_Max/2 y +Shock_Max. "
        "Cambie Shock_Max y toda la matriz se recalcula."
    )
    ws["H4"].font = FONT_SUB
    ws["H4"].alignment = ALIGN_WRAP
    ws.merge_cells("H4:H7")

    ws["A6"] = "MtM(Δs, Δc) = Σ ε·(K − (F + Δs)·(1 + Δc))·N·DF     con Δs = Spot × %spot, Δc = %curva"
    ws["A6"].font = Font(name="Calibri", size=10, italic=True, color=C_HEADER)
    ws.merge_cells("A6:F6")

    ws.merge_cells("B8:F8")
    c = ws["B8"]
    c.value = "Desplazamiento de la curva forward (%)"
    c.fill = FILL_HEADER_SOFT
    c.font = FONT_HEADER
    c.alignment = ALIGN_C

    _hdr(ws, 9, 1, "Desplazamiento del spot (%)")
    cols = ["B", "C", "D", "E", "F"]
    factores = ["-Shock_Max", "-Shock_Max/2", "0", "Shock_Max/2", "Shock_Max"]
    for col, fac in zip(cols, factores):
        c = ws[f"{col}9"]
        c.value = f"={fac}" if fac != "0" else "=0"
        c.fill = FILL_HEADER
        c.font = FONT_HEADER
        c.alignment = ALIGN_C
        c.border = BORDER_ALL
        c.number_format = FMT_PCTLBL

    # filas: spot de mayor a menor (spot más alto arriba, como el motor)
    filas_fac = ["Shock_Max", "Shock_Max/2", "0", "-Shock_Max/2", "-Shock_Max"]
    for i, fac in enumerate(filas_fac):
        r = 10 + i
        c = ws[f"A{r}"]
        c.value = f"={fac}" if fac != "0" else "=0"
        c.fill = FILL_HEADER
        c.font = FONT_HEADER
        c.alignment = ALIGN_C
        c.border = BORDER_ALL
        c.number_format = FMT_PCTLBL
        for col in cols:
            cell = ws[f"{col}{r}"]
            cell.value = (
                f"=SUMPRODUCT({rZ},({rAA}-({rAB}+Spot_Valorizacion*$A{r}/100)*(1+{col}$9/100)))"
            )
            cell.number_format = FMT_CLP
            cell.border = BORDER_ALL
            cell.font = FONT_BASE
            cell.alignment = ALIGN_R

    ws.conditional_formatting.add(
        "B10:F14",
        ColorScaleRule(
            start_type="min", start_color="F8696B",
            mid_type="num", mid_value=0, mid_color="FFFFFF",
            end_type="max", end_color="63BE7B",
        ),
    )

    ws["A16"] = "MtM base (sin shock)"
    ws["A16"].font = FONT_BOLD
    _calc(ws, "B16", "=D12", FMT_CLP, bold=True)
    ws["A17"] = "Peor escenario"
    ws["A17"].font = FONT_BOLD
    _calc(ws, "B17", "=MIN(B10:F14)", FMT_CLP, bold=True)
    ws["A18"] = "Mejor escenario"
    ws["A18"].font = FONT_BOLD
    _calc(ws, "B18", "=MAX(B10:F14)", FMT_CLP, bold=True)
    ws["A19"] = "Pérdida máxima frente a la base"
    ws["A19"].font = FONT_BOLD
    _calc(ws, "B19", "=B17-B16", FMT_CLP, bold=True)

    ws["H10"] = (
        "El producto es lineal en el forward, de modo que la revalorización completa admite forma "
        "cerrada: desplazar la curva de outrights en Δs y escalarla en (1+Δc) equivale, por ser la "
        "interpolación lineal, a aplicar la misma transformación afín al forward ya interpolado. "
        "Por eso cada celda es un único SUMPRODUCT y no necesita reinterpolar la curva."
    )
    ws["H10"].font = FONT_SUB
    ws["H10"].alignment = ALIGN_WRAP
    ws.merge_cells("H10:H16")


# ──────────────────────────────────────────────────────────────────────
# Hoja: Griegas
# ──────────────────────────────────────────────────────────────────────

def _sheet_griegas(wb: Workbook, data: WorkbookData) -> None:
    ws = wb[SH_GRI]
    ws.sheet_view.showGridLines = False
    nrows = data.n_rows
    first, last = VAL_FIRST_ROW, VAL_FIRST_ROW + nrows - 1
    tot = last + 2
    VAL = q(SH_VAL)

    _widths(ws, {"A": 30, "B": 20, "C": 16, "D": 18, "E": 18, "F": 18, "G": 16, "H": 4, "I": 62})
    _title(ws, "Griegas y exposición",
           "Sensibilidades agregadas de la cartera y desglose por contraparte.")

    _hdr(ws, 4, 1, "Medida")
    _hdr(ws, 4, 2, "Valor")
    _hdr(ws, 4, 4, "Interpretación")
    ws.merge_cells("D4:G4")

    medidas = [
        ("MtM total (CLP)", f"={VAL}!$O${tot}", FMT_CLP,
         "Valor de mercado de la cartera a la fecha de valorización."),
        ("Componente spot (CLP)", f"={VAL}!$P${tot}", FMT_CLP,
         "Parte del MtM atribuible al movimiento del spot desde que se pactó cada operación."),
        ("Puntos forward (CLP)", f"={VAL}!$Q${tot}", FMT_CLP,
         "Resto del MtM: diferencial de tasas y paso del tiempo."),
        ("Delta total (CLP por 1 CLP)", f"={VAL}!$R${tot}", FMT_CLP,
         "Variación del MtM ante un alza de 1 peso en el spot, con puntos forward constantes."),
        ("Delta por 1% de spot (CLP)", f"={VAL}!$R${tot}*Spot_Valorizacion/100", FMT_CLP,
         "La misma delta expresada para un movimiento de 1% del tipo de cambio."),
        ("DV01 total (CLP)", f"={VAL}!$S${tot}", FMT_CLP2,
         "Variación del MtM ante un alza paralela de 1 punto base en la curva de descuento."),
        ("Theta 1 día (CLP)", f"={VAL}!$AJ${tot}", FMT_CLP2,
         "Variación del MtM por un día de paso del tiempo con las curvas congeladas."),
        ("Gamma", "=0", "0",
         "Exactamente cero: el pago de un forward es lineal en el subyacente."),
        ("Nocional total", f"={VAL}!$K${tot}", FMT_CLP,
         "Suma de nocionales en moneda base."),
        ("Contratos valorizados", f"={VAL}!$D${tot}", "0",
         "Filas con nocional cargado."),
    ]
    r = 5
    for nombre, formula, fmt, expl in medidas:
        _label(ws, f"A{r}", nombre)
        ws[f"A{r}"].border = BORDER_ALL
        _calc(ws, f"B{r}", formula, fmt, bold=True)
        e = ws[f"D{r}"]
        e.value = expl
        e.font = FONT_SUB
        e.alignment = ALIGN_L
        ws.merge_cells(f"D{r}:G{r}")
        r += 1

    # ── desglose por contraparte ─────────────────────────────────────
    r += 1
    ws[f"A{r}"] = "DESGLOSE POR CONTRAPARTE"
    ws[f"A{r}"].font = Font(name="Calibri", size=12, bold=True, color=C_HEADER)
    r += 1
    head_row = r
    encabezados = ["Contraparte", "Nº contratos", "Nocional", "MtM (CLP)",
                   "Componente spot", "Delta", "DV01"]
    for j, h in enumerate(encabezados, start=1):
        _hdr(ws, head_row, j, h)
    ws.row_dimensions[head_row].height = 28

    contrapartes: list[str] = []
    for c in data.contracts:
        nombre = (c.counterparty or "").strip()
        if nombre and nombre not in contrapartes:
            contrapartes.append(nombre)
    slots = max(25, len(contrapartes))

    rango_cp = f"{VAL}!$B${first}:$B${last}"
    sum_cols = {"C": "K", "D": "O", "E": "P", "F": "R", "G": "S"}
    for i in range(slots):
        rr = head_row + 1 + i
        cell = ws[f"A{rr}"]
        cell.value = contrapartes[i] if i < len(contrapartes) else None
        cell.fill = FILL_INPUT
        cell.border = BORDER_ALL
        cell.font = FONT_BASE
        cell.alignment = ALIGN_L
        b = ws[f"B{rr}"]
        b.value = f'=IF($A{rr}="","",COUNTIF({rango_cp},$A{rr}))'
        b.number_format = "0"
        b.fill = FILL_CALC
        b.border = BORDER_ALL
        b.font = FONT_BASE
        b.alignment = ALIGN_C
        for col, vcol in sum_cols.items():
            c2 = ws[f"{col}{rr}"]
            c2.value = (
                f'=IF($A{rr}="","",SUMIF({rango_cp},$A{rr},'
                f"{VAL}!${vcol}${first}:${vcol}${last}))"
            )
            c2.number_format = FMT_CLP2 if col == "G" else FMT_CLP
            c2.fill = FILL_CALC
            c2.border = BORDER_ALL
            c2.font = FONT_BASE
            c2.alignment = ALIGN_R

    tr = head_row + 1 + slots
    ws[f"A{tr}"] = "TOTAL"
    ws[f"A{tr}"].font = FONT_BOLD
    for col in "ABCDEFG":
        ws[f"{col}{tr}"].fill = FILL_TOTAL
        ws[f"{col}{tr}"].border = BORDER_TOP
    for col in ("B", "C", "D", "E", "F", "G"):
        c3 = ws[f"{col}{tr}"]
        c3.value = f"=SUM({col}{head_row + 1}:{col}{tr - 1})"
        c3.number_format = "0" if col == "B" else (FMT_CLP2 if col == "G" else FMT_CLP)
        c3.font = FONT_BOLD
        c3.alignment = ALIGN_R

    ws[f"I{head_row}"] = (
        "Los nombres de contraparte son celdas de entrada: el desglose usa SUMAR.SI sobre la columna "
        "Contraparte de la hoja Valorización, de modo que si agrega operaciones de una contraparte "
        "nueva basta escribir su nombre en una fila libre de esta tabla."
    )
    ws[f"I{head_row}"].font = FONT_SUB
    ws[f"I{head_row}"].alignment = ALIGN_WRAP
    ws.merge_cells(f"I{head_row}:I{head_row + 6}")

    ws.conditional_formatting.add(
        f"D{head_row + 1}:G{tr}",
        CellIsRule(operator="lessThan", formula=["0"], font=Font(color=C_NEG)),
    )


# ──────────────────────────────────────────────────────────────────────
# Hoja: Reconciliación
# ──────────────────────────────────────────────────────────────────────

REC_FIRST_ROW = 6


def _sheet_reconciliacion(wb: Workbook, data: WorkbookData) -> None:
    ws = wb[SH_REC]
    ws.sheet_view.showGridLines = False
    ref = data.referencia or []
    nrows = data.n_rows
    first = REC_FIRST_ROW
    last = first + nrows - 1
    VAL = q(SH_VAL)

    _title(ws, "Reconciliación Excel contra motor Python",
           "Las columnas 'motor' son valores calculados por core.valuation al generar el archivo; "
           "las columnas 'Excel' son fórmulas vivas de la hoja Valorización.")

    _label(ws, "A4", "RESULTADO GLOBAL")
    ws["A4"].font = Font(name="Calibri", size=12, bold=True, color=C_HEADER)
    est = ws["C4"]
    est.value = (
        f'=IF(COUNT($M${first}:$M${last})=0,"SIN DATOS",'
        f'IF(MAX($M${first}:$M${last},$P${first}:$P${last},$H${first}:$H${last},'
        f'$K${first}:$K${last})<=0.01,"CUADRA","NO CUADRA"))'
    )
    est.font = Font(name="Calibri", size=14, bold=True, color=C_OK)
    est.fill = FILL_CALC
    est.border = BORDER_ALL
    est.alignment = ALIGN_C
    ws.merge_cells("C4:D4")
    ws["E4"] = "Tolerancia: 0,01 CLP en MtM y componente spot; 1e-2 en forward y tasa."
    ws["E4"].font = FONT_SUB

    ws.conditional_formatting.add(
        "C4",
        CellIsRule(operator="equal", formula=['"NO CUADRA"'],
                   font=Font(color="FFFFFF", bold=True),
                   fill=PatternFill("solid", fgColor=C_NEG)),
    )

    cols = [
        ("A", "Folio", 12, None),
        ("B", "Contraparte", 18, None),
        ("C", "F mercado Excel", 16, FMT_PRICE),
        ("D", "F mercado motor", 16, FMT_PRICE),
        ("E", "Tasa Excel (%)", 15, FMT_RATE),
        ("F", "Tasa motor (%)", 15, FMT_RATE),
        ("G", "DF Excel", 15, FMT_DF),
        ("H", "|Δ| tasa+fwd", 14, "0.00000000"),
        ("I", "MtM Excel", 18, FMT_CLP2),
        ("J", "MtM motor", 18, FMT_CLP2),
        ("K", "|Δ| MtM", 14, "0.00000000"),
        ("L", "Δ relativo MtM", 14, "0.00E+00"),
        ("M", "|Δ| MtM (control)", 16, "0.00000000"),
        ("N", "Comp. spot Excel", 18, FMT_CLP2),
        ("O", "Comp. spot motor", 18, FMT_CLP2),
        ("P", "|Δ| comp. spot", 15, "0.00000000"),
        ("Q", "Estado", 14, None),
    ]
    for col, nombre, w, _f in cols:
        _hdr(ws, 5, ws[f"{col}1"].column, nombre, w)
    ws.row_dimensions[5].height = 30

    for i in range(nrows):
        r = first + i
        rv = VAL_FIRST_ROW + i
        item = ref[i] if i < len(ref) else {}
        tiene = bool(item)

        ws[f"A{r}"] = f'=IF({VAL}!$A{rv}="","",{VAL}!$A{rv})'
        ws[f"B{r}"] = f'=IF({VAL}!$B{rv}="","",{VAL}!$B{rv})'
        ws[f"C{r}"] = f'=IF({VAL}!$H{rv}="","",{VAL}!$H{rv})'
        ws[f"D{r}"] = item.get("fwd_mkt") if tiene else None
        ws[f"E{r}"] = f'=IF({VAL}!$I{rv}="","",{VAL}!$I{rv})'
        ws[f"F{r}"] = item.get("disc_rate") if tiene else None
        ws[f"G{r}"] = f'=IF({VAL}!$J{rv}="","",{VAL}!$J{rv})'
        ws[f"H{r}"] = (
            f'=IF(OR($C{r}="",$D{r}=""),"",MAX(ABS($C{r}-$D{r}),ABS($E{r}-$F{r})))'
        )
        ws[f"I{r}"] = f'=IF({VAL}!$O{rv}="","",{VAL}!$O{rv})'
        ws[f"J{r}"] = item.get("mtm") if tiene else None
        ws[f"K{r}"] = f'=IF(OR($I{r}="",$J{r}=""),"",ABS($I{r}-$J{r}))'
        ws[f"L{r}"] = f'=IF(OR($K{r}="",N($J{r})=0),"",$K{r}/ABS($J{r}))'
        ws[f"M{r}"] = f'=IF($K{r}="","",$K{r})'
        ws[f"N{r}"] = f'=IF({VAL}!$P{rv}="","",{VAL}!$P{rv})'
        ws[f"O{r}"] = item.get("spot_component") if tiene else None
        ws[f"P{r}"] = f'=IF(OR($N{r}="",$O{r}=""),"",ABS($N{r}-$O{r}))'
        ws[f"Q{r}"] = (
            f'=IF($K{r}="","",IF(AND($K{r}<=0.01,$P{r}<=0.01,$H{r}<=0.01),"OK","REVISAR"))'
        )

        for col, _n, _w, fmt in cols:
            cell = ws[f"{col}{r}"]
            cell.border = BORDER_ALL
            cell.font = FONT_BASE
            if fmt:
                cell.number_format = fmt
            cell.alignment = ALIGN_L if col in ("A", "B", "Q") else ALIGN_R
            cell.fill = FILL_INPUT if col in ("D", "F", "J", "O") else FILL_CALC

    for col in ("H", "K", "M", "P"):
        ws.conditional_formatting.add(
            f"{col}{first}:{col}{last}",
            CellIsRule(operator="greaterThan", formula=["0.01"],
                       font=Font(color="FFFFFF", bold=True),
                       fill=PatternFill("solid", fgColor=C_NEG)),
        )
    ws.conditional_formatting.add(
        f"Q{first}:Q{last}",
        CellIsRule(operator="equal", formula=['"REVISAR"'],
                   font=Font(color="FFFFFF", bold=True),
                   fill=PatternFill("solid", fgColor=C_NEG)),
    )
    ws.conditional_formatting.add(
        f"Q{first}:Q{last}",
        CellIsRule(operator="equal", formula=['"OK"'], font=Font(color=C_OK, bold=True)),
    )

    ws.auto_filter.ref = f"A5:Q{last}"
    ws.freeze_panes = "C6"


# ──────────────────────────────────────────────────────────────────────
# Hoja: Metodología
# ──────────────────────────────────────────────────────────────────────

_METODOLOGIA = [
    ("H1", "Metodología de valorización"),
    ("P", "Este libro replica, con fórmulas nativas de Excel, el motor de valorización "
          "implementado en core/valuation.py. No hay macros, VBA ni funciones definidas por el "
          "usuario: el archivo se abre y recalcula en cualquier aplicación de planilla."),

    ("H2", "1. Convención de signos"),
    ("F", "ε = +1 si la operación es una VENTA de la moneda base"),
    ("F", "ε = −1 si la operación es una COMPRA de la moneda base"),
    ("P", "El vendedor gana cuando el mercado cae por debajo del precio pactado; de ahí el signo. "
          "En la hoja Valorización la columna ε se deriva del campo Operación de la hoja Contratos."),

    ("H2", "2. Valor de mercado"),
    ("F", "MtM = ε · (K − F_mercado) · N · DF"),
    ("P", "K es el precio forward pactado, F_mercado el forward de mercado al plazo residual, "
          "N el nocional en moneda base y DF el factor de descuento en moneda de cotización. "
          "El resultado queda expresado en moneda de cotización (CLP)."),

    ("H2", "3. Descomposición del resultado"),
    ("F", "Componente spot = ε · (S₀ − S_valorización) · N · DF"),
    ("F", "Puntos forward   = MtM − Componente spot"),
    ("P", "S₀ es el tipo de cambio spot vigente el día en que se pactó la operación. La descomposición "
          "separa el resultado atribuible al movimiento del tipo de cambio del atribuible al "
          "diferencial de tasas y al paso del tiempo. Si S₀ no está informado, el componente spot se "
          "fija en cero y la fila se marca con una alerta: la descomposición no sería confiable."),

    ("H2", "4. Plazo y fracción de año"),
    ("F", "Días a vencimiento = Fecha de vencimiento − Fecha de valorización   (días corridos)"),
    ("F", "t = Días a vencimiento / Base anual"),
    ("P", "La base anual es 360 por defecto (ACT/360, convención del mercado local). Cambiándola a 365 "
          "el libro pasa a ACT/365. Las fechas no se ajustan a día hábil: se usa la fecha de "
          "vencimiento tal cual está cargada, que es la convención 'Exacto' del motor."),

    ("H2", "5. Factor de descuento"),
    ("F", "DF = (1 + tasa/100) ^ (−t)"),
    ("P", "Capitalización compuesta anual sobre la tasa cero expresada en porcentaje, que es la "
          "convención del mercado CLP. La tasa se interpola en la curva de descuento al mismo plazo "
          "residual del contrato."),

    ("H2", "6. Interpolación y extrapolación de curvas"),
    ("P", "Ambas curvas se interpolan LINEALMENTE en días al vencimiento. Sea x el plazo consultado y "
          "(x₁,y₁)…(xₙ,yₙ) los nodos ordenados en forma ascendente:"),
    ("F", "k = MAX(1; MIN(n−1; CONTAR.SI(dias; \"<=\" & x)))"),
    ("F", "y = y_k + (x − x_k) · (y_{k+1} − y_k) / (x_{k+1} − x_k)"),
    ("P", "La misma expresión cubre los tres regímenes. Si x es menor que el primer nodo, CONTAR.SI "
          "devuelve 0, k queda en 1 y la pendiente de los dos primeros nodos se prolonga hacia atrás. "
          "Si x es mayor o igual que el último nodo, CONTAR.SI devuelve n, k queda en n−1 y se "
          "prolonga la pendiente de los dos últimos. Dentro del rango, k es el índice del nodo "
          "inferior y la fórmula interpola. Se usa CONTAR.SI y no COINCIDIR porque el bloque de "
          "curvas reserva 60 celdas y COINCIDIR con coincidencia aproximada exige un rango "
          "estrictamente ordenado sin celdas vacías."),
    ("P", "La extrapolación PLANA se obtiene sin una segunda fórmula: la columna auxiliar 'plazo "
          "efectivo' acota x al intervalo [x₁, xₙ] cuando el parámetro Metodo_Extrapolacion vale "
          "'Plana'. Con 'Lineal' el plazo pasa sin modificación y la fórmula de pendiente extrapola."),
    ("P", "Advertencia metodológica: la planilla Cordada original y este libro usan extrapolación "
          "LINEAL. La primera versión de la aplicación web usaba extrapolación plana y por eso no "
          "reproducía los MtM de la planilla en los contratos de plazo corto, cuyos días residuales "
          "caen por debajo del primer nodo de la curva de descuento (92 días)."),

    ("H2", "7. Sensibilidades"),
    ("F", "Delta = −ε · N · DF          (variación del MtM por 1 CLP de alza del spot)"),
    ("F", "DV01  = MtM(tasa + 1 pb) − MtM"),
    ("F", "Theta = MtM(t+1 día, curvas congeladas) − MtM"),
    ("F", "Gamma = 0                    (el pago es lineal en el subyacente)"),
    ("P", "La delta supone que la curva de outrights se desplaza uno a uno con el spot manteniendo "
          "constantes los puntos forward. El DV01 aplica un desplazamiento paralelo de 1 punto base "
          "a la curva de descuento y revaloriza. La theta adelanta un día la fecha de valorización, "
          "reinterpolando ambas curvas al nuevo plazo residual."),

    ("H2", "8. Matriz de sensibilidad"),
    ("F", "MtM(Δs, Δc) = Σ ε · (K − (F + Δs) · (1 + Δc)) · N · DF"),
    ("P", "Δs es el desplazamiento absoluto del spot (porcentaje del spot vigente) y Δc un factor "
          "multiplicativo sobre la curva de outrights ya desplazada. Como la interpolación es lineal, "
          "aplicar una transformación afín a los nodos equivale a aplicarla al forward interpolado: "
          "cada celda de la matriz es una revalorización completa, no una aproximación de primer "
          "orden, y se resuelve con un único SUMAPRODUCTO. El factor de descuento no depende del "
          "spot ni del nivel de la curva forward, por lo que permanece fijo dentro de la matriz."),

    ("H2", "9. Alertas"),
    ("P", "La columna Alertas de la hoja Valorización marca: contratos vencidos o que vencen el mismo "
          "día de valorización, ausencia de spot al inicio, nocional o precio pactado no positivos, y "
          "plazos que caen fuera del rango de nodos de cualquiera de las dos curvas (forward "
          "extrapolado o descuento extrapolado). Una alerta no excluye la fila de los totales: el "
          "criterio queda a juicio del analista."),

    ("H2", "10. Diferencias conocidas frente al motor Python"),
    ("P", "· El motor excluye de los totales los contratos con error fatal (nocional no positivo, "
          "precio pactado no positivo, contrato vencido). La fila TOTAL de este libro suma todas las "
          "filas con nocional; las filas problemáticas quedan señaladas en la columna Alertas."),
    ("P", "· El motor puede ajustar el vencimiento a día hábil según el calendario chileno. Este "
          "libro usa la convención 'Exacto' (fecha tal cual), que es la de la planilla de referencia."),
    ("P", "· CVA y DVA no se calculan en el libro: dependen de un modelo de exposición esperada que "
          "no admite una expresión cerrada razonable en Excel. El MtM ajustado por crédito se "
          "consulta en la aplicación."),

    ("H2", "11. Cómo usar el libro"),
    ("P", "Edite únicamente las celdas azules: los parámetros, los nodos de las dos curvas y la tabla "
          "de contratos. Todo lo demás se recalcula solo. Para agregar operaciones, escriba en la "
          "primera fila libre de la hoja Contratos; la hoja Valorización ya tiene las fórmulas "
          "preparadas para todas las filas. Para agregar nodos de curva, continúe el bloque hacia "
          "abajo sin dejar filas vacías intermedias y manteniendo el orden ascendente de plazo."),
]


def _sheet_metodologia(wb: Workbook, data: WorkbookData) -> None:
    ws = wb[SH_MET]
    ws.sheet_view.showGridLines = False
    _widths(ws, {"A": 3, "B": 118})

    ws["B1"] = "Metodología"
    ws["B1"].font = FONT_TITLE
    r = 3
    for tipo, texto in _METODOLOGIA:
        c = ws.cell(row=r, column=2, value=texto)
        if tipo == "H1":
            c.font = Font(name="Calibri", size=14, bold=True, color=C_HEADER)
            ws.row_dimensions[r].height = 22
        elif tipo == "H2":
            c.font = Font(name="Calibri", size=12, bold=True, color=C_HEADER)
            ws.row_dimensions[r].height = 24
        elif tipo == "F":
            c.font = Font(name="Consolas", size=10, color="1F3864")
            c.fill = FILL_CALC
            c.border = BORDER_ALL
            ws.row_dimensions[r].height = 18
        else:
            c.font = FONT_BASE
            c.alignment = ALIGN_WRAP
            ws.row_dimensions[r].height = max(15, 14 * (len(texto) // 118 + 1))
        if tipo != "F":
            c.alignment = ALIGN_WRAP
        r += 1

    ws.cell(row=r + 1, column=2,
            value=f"Modelo versión {VERSION} — generado el "
                  f"{datetime.now().strftime('%d-%m-%Y %H:%M')}").font = FONT_SUB


# ──────────────────────────────────────────────────────────────────────
# Ensamblaje
# ──────────────────────────────────────────────────────────────────────

_BUILDERS = {
    SH_PAR: _sheet_parametros,
    SH_CUR: _sheet_curvas,
    SH_CON: _sheet_contratos,
    SH_VAL: _sheet_valorizacion,
    SH_SEN: _sheet_sensibilidad,
    SH_GRI: _sheet_griegas,
    SH_REC: _sheet_reconciliacion,
    SH_MET: _sheet_metodologia,
}


def build_workbook_object(data: WorkbookData, hojas: tuple[str, ...] | None = None) -> Workbook:
    """Construye el libro completo y lo devuelve como objeto openpyxl."""
    hojas = tuple(hojas) if hojas else ALL_SHEETS
    hojas = tuple(h for h in ALL_SHEETS if h in hojas)   # orden canónico

    if data.referencia is None:
        try:
            data.referencia = compute_reference(data)
        except Exception:
            data.referencia = []

    wb = Workbook()
    wb.remove(wb.active)
    for nombre in hojas:
        ws = wb.create_sheet(nombre)
        ws.sheet_properties.tabColor = C_HEADER

    if SH_POR in hojas:
        _sheet_portada(wb, data, hojas)
    for nombre in hojas:
        builder = _BUILDERS.get(nombre)
        if builder:
            builder(wb, data)

    wb.calculation.fullCalcOnLoad = True
    props = wb.properties
    props.title = data.titulo
    props.subject = "Valorización de forwards FX"
    props.creator = "Valorizador de Forwards"
    props.description = f"Modelo con fórmulas vivas, versión {VERSION}."
    wb.active = 0
    return wb


def build_workbook_bytes(data: WorkbookData, hojas: tuple[str, ...] | None = None) -> io.BytesIO:
    """Igual que `build_workbook_object` pero devuelve un `io.BytesIO` listo para servir."""
    wb = build_workbook_object(data, hojas)
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


def save_workbook(data: WorkbookData, ruta: str, hojas: tuple[str, ...] | None = None) -> str:
    wb = build_workbook_object(data, hojas)
    wb.save(ruta)
    return ruta


# ──────────────────────────────────────────────────────────────────────
# Datos de demostración (caso Cordada 31-05-2026)
# ──────────────────────────────────────────────────────────────────────

def demo_data() -> WorkbookData:
    """
    Caso de referencia del libro Cordada 31-05-2026.

    Los nodos y contratos se leen de
    `valorizador/management/commands/cargar_demo.py` cuando ese módulo se puede
    importar; si Django no está configurado se usa la copia local, que es
    idéntica y está cubierta por `verificar_excel`.
    """
    fwd, desc, contratos, spot, fecha = _demo_constantes()
    return WorkbookData(
        valuation_date=fecha,
        spot=float(spot),
        fwd_nodes=[(float(d), float(v)) for d, v in fwd],
        desc_nodes=[(float(d), float(v)) for d, v in desc],
        contracts=[
            ContractRow(
                folio=folio,
                counterparty=cp,
                cartera="Cordada",
                side="Venta",
                modality="Compensacion",
                base_ccy="USD",
                notional=float(nocional),
                fwd_price=float(precio),
                spot_inicio=float(spot_ini),
                start_date=None,
                maturity_date=vcto,
            )
            for folio, cp, vcto, nocional, spot_ini, precio in contratos
        ],
        base_anual=360,
        extrap="Lineal",
        day_count="ACT/360",
        compounding="Compuesta",
        label="Cordada",
        fuente="Libro Cordada 31-05-2026",
        titulo="Valorizador de Forwards FX USD/CLP — Cartera Cordada",
    )


_DEMO_FWD = [
    (1, 892.21), (2, 892.205), (8, 892.19), (15, 892.13), (22, 892.105),
    (31, 892.06), (62, 892.03),
]
_DEMO_DESC = [
    (92, 3.48231), (183, 3.61177), (271, 3.70649), (365, 3.78017),
    (731, 3.98414), (1096, 4.24534), (1461, 4.42915),
]
_DEMO_CONTRATOS = [
    ("756929", "BTG Pactual", date(2026, 7, 7), 1_000_000, 887.71, 886.94),
    ("118039", "Bice", date(2026, 7, 13), 2_000_000, 894.25, 893.35),
    ("116845", "Bice", date(2026, 6, 12), 2_000_000, 890.33, 889.98),
]
_DEMO_SPOT = 892.89
_DEMO_FECHA = date(2026, 5, 31)


def _demo_constantes():
    """Intenta leer las constantes del comando de Django; si no, usa la copia local."""
    try:  # pragma: no cover - depende del entorno
        import os

        os.environ.setdefault("DJANGO_SETTINGS_MODULE", "config.settings")
        import django

        django.setup()
        from valorizador.management.commands import cargar_demo as cd

        return (
            list(cd.FWD_NODOS),
            list(cd.DESC_NODOS),
            list(cd.CONTRATOS),
            float(cd.SPOT_VALORIZACION),
            cd.FECHA_VALORIZACION,
        )
    except Exception:
        return (
            list(_DEMO_FWD), list(_DEMO_DESC), list(_DEMO_CONTRATOS),
            _DEMO_SPOT, _DEMO_FECHA,
        )

"""
Lector del libro `CalculadoraForward Cordada_*.xlsm`.

El cargador original hardcodeaba nombres de hoja y posiciones de celda
("Forwards Cordada 31", 'B5', 'C1', filas 2..29 de CURVE MASTER) y fallaba en
silencio si el libro del mes siguiente cambiaba de nombre. Acá:

* Las hojas se localizan por patrón, no por nombre exacto.
* Las columnas de `FWD Vigentes` se localizan por encabezado.
* El rango de `CURVE MASTER` se lee hasta que se acaban los datos.
* `spot_inicio` se toma de la columna "Tipo de Cambio al Inicio" de la hoja de
  valorización cuando existe. El cargador original asignaba a todos los
  contratos el spot **de hoy** como si fuera el spot de la fecha de pacto, lo
  que dejaba el componente spot en cero por construcción.
* Devuelve un informe de lo que encontró en vez de un mensaje genérico.
"""

from __future__ import annotations

import re
from datetime import date, datetime

__all__ = ["CordadaWorkbook"]


def _as_date(v):
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    return None


def _as_float(v):
    try:
        f = float(v)
        return f if f == f else None
    except (TypeError, ValueError):
        return None


class CordadaWorkbook:
    """Extrae curvas, spot, fecha y contratos de un libro Cordada."""

    def __init__(self, path_or_file):
        import openpyxl

        self.wb = openpyxl.load_workbook(path_or_file, data_only=True)
        self.warnings: list[str] = []

    # -- localización de hojas ----------------------------------------

    def _find_sheet(self, *patterns: str):
        for pat in patterns:
            rx = re.compile(pat, re.IGNORECASE)
            for name in self.wb.sheetnames:
                if rx.search(name):
                    return self.wb[name]
        return None

    # -- extracción ---------------------------------------------------

    def valuation_date_and_spot(self) -> tuple[date | None, float | None]:
        """Fecha de valorización y tipo de cambio de esa fecha."""
        ws = self._find_sheet(r'forwards\s+cordada')
        fecha = spot = None

        if ws is not None:
            # La fecha suele estar en la primera fila, junto a la etiqueta "Fecha".
            for row in ws.iter_rows(min_row=1, max_row=4, values_only=True):
                for i, cell in enumerate(row):
                    if isinstance(cell, str) and cell.strip().lower() == 'fecha':
                        for nxt in row[i + 1:i + 3]:
                            fecha = fecha or _as_date(nxt)
                    fecha = fecha or _as_date(cell)
                if fecha:
                    break

            # El spot es el primer valor de la columna "Tipo de Cambio Fecha ...".
            header_row, col_idx = self._locate_header(ws, r'tipo de cambio fecha')
            if header_row and col_idx is not None:
                for row in ws.iter_rows(min_row=header_row + 1, max_row=header_row + 30,
                                        values_only=True):
                    v = _as_float(row[col_idx]) if col_idx < len(row) else None
                    if v and v > 0:
                        spot = round(v, 4)
                        break

        if fecha is None:
            ws_datos = self._find_sheet(r'^datos$')
            if ws_datos is not None:
                for row in ws_datos.iter_rows(min_row=1, max_row=20, values_only=True):
                    for cell in row:
                        fecha = fecha or _as_date(cell)
                    if fecha:
                        break

        if fecha is None:
            self.warnings.append(
                'No se encontró la fecha de valorización en el libro; se usará la fecha de hoy.'
            )
        if spot is None:
            self.warnings.append(
                'No se encontró el tipo de cambio de la fecha de valorización.'
            )
        return fecha, spot

    @staticmethod
    def _locate_header(ws, pattern: str, max_row: int = 12):
        rx = re.compile(pattern, re.IGNORECASE)
        for r_i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_row, values_only=True), 1):
            for c_i, cell in enumerate(row):
                if isinstance(cell, str) and rx.search(cell):
                    return r_i, c_i
        return None, None

    def curves(self) -> dict[str, list[dict]]:
        """
        Nodos de las curvas de `CURVE MASTER`.

        La hoja tiene pares de columnas (dia_X, c_X). Se leen todos los pares
        presentes, no sólo los dos primeros.
        """
        ws = self._find_sheet(r'curve\s*master')
        if ws is None:
            self.warnings.append("No se encontró la hoja 'CURVE MASTER'.")
            return {}

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return {}

        headers = [str(h).strip() if h is not None else '' for h in rows[0]]
        out: dict[str, list[dict]] = {}

        for i, h in enumerate(headers):
            m = re.match(r'^dia[_\s]*(.+)$', h, re.IGNORECASE)
            if not m or i + 1 >= len(headers):
                continue
            curve_name = m.group(1).strip().upper()
            puntos = []
            for row in rows[1:]:
                if i + 1 >= len(row):
                    continue
                d, v = _as_float(row[i]), _as_float(row[i + 1])
                if d is None or v is None:
                    continue
                puntos.append({'tenor_days': int(round(d)), 'value': float(v)})
            if puntos:
                out[curve_name] = puntos

        if not out:
            self.warnings.append("'CURVE MASTER' no tenía pares (dia_X, c_X) reconocibles.")
        return out

    def contracts(self) -> list[dict]:
        """Contratos vigentes desde la hoja `FWD Vigentes ...`."""
        ws = self._find_sheet(r'fwd\s+vigentes', r'vigentes')
        if ws is None:
            self.warnings.append("No se encontró la hoja 'FWD Vigentes'.")
            return []

        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return []

        headers = [str(h).strip().lower() if h is not None else '' for h in rows[0]]

        def col(*names, default=None):
            for n in names:
                for i, h in enumerate(headers):
                    if h == n:
                        return i
            for n in names:
                for i, h in enumerate(headers):
                    if n in h:
                        return i
            return default

        i_cp = col('contraparte', 'counterparty')
        i_folio = col('folio', 'ref')
        i_estado = col('estado')
        i_oper = col('operación', 'operacion', 'side')
        i_moda = col('modalidad')
        i_ini = col('fecha inicio')
        i_vcto = col('fecha vcto', 'vcto', 'vencim')
        i_monto = col('monto')
        i_mon = col('moneda')
        i_precio = col('precio fwd', 'precio')

        # Mapa de spot al inicio desde la hoja de valorización, por folio.
        spot_por_folio = self._spot_inicio_por_folio()

        out = []
        for row in rows[1:]:
            if i_cp is None or i_vcto is None or i_monto is None:
                break
            cp = row[i_cp] if i_cp < len(row) else None
            vcto = _as_date(row[i_vcto]) if i_vcto < len(row) else None
            monto = _as_float(row[i_monto]) if i_monto < len(row) else None
            if not cp or vcto is None or not monto:
                continue
            if i_estado is not None and i_estado < len(row):
                estado = str(row[i_estado] or '').strip().lower()
                if estado and not estado.startswith('vigente'):
                    continue

            folio = str(row[i_folio]).strip() if i_folio is not None and row[i_folio] else ''
            precio = _as_float(row[i_precio]) if i_precio is not None and i_precio < len(row) else None
            side = str(row[i_oper]).strip() if i_oper is not None and row[i_oper] else 'Venta'
            side = 'Compra' if side.lower().startswith('c') else 'Venta'

            out.append({
                'counterparty': str(cp).strip(),
                'folio': folio,
                'side': side,
                'modality': str(row[i_moda]).strip() if i_moda is not None and row[i_moda] else 'Compensacion',
                'base_ccy': (str(row[i_mon]).strip().upper()[:3] if i_mon is not None and row[i_mon] else 'USD'),
                'quote_ccy': 'CLP',
                'notional': round(monto, 2),
                'fwd_price': round(precio, 4) if precio else 0.0,
                'spot_inicio': spot_por_folio.get(folio, 0.0),
                'start_date': _as_date(row[i_ini]) if i_ini is not None and i_ini < len(row) else None,
                'maturity_date': vcto,
                'status': 'Vigente',
            })

        if not out:
            self.warnings.append('La hoja de contratos vigentes no tenía filas utilizables.')
        return out

    def _spot_inicio_por_folio(self) -> dict[str, float]:
        """Lee la columna 'Tipo de Cambio al Inicio' de la hoja de valorización."""
        ws = self._find_sheet(r'forwards\s+cordada')
        if ws is None:
            return {}

        header_row, i_spot = self._locate_header(ws, r'tipo de cambio al inicio')
        _, i_ref = self._locate_header(ws, r'^ref$')
        if header_row is None or i_spot is None or i_ref is None:
            self.warnings.append(
                'No se pudo mapear el tipo de cambio al inicio por folio; el componente '
                'spot quedará en cero hasta que lo completes manualmente.'
            )
            return {}

        out: dict[str, float] = {}
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if i_ref >= len(row) or i_spot >= len(row):
                continue
            ref, spot = row[i_ref], _as_float(row[i_spot])
            if ref is None or spot is None:
                continue
            out[str(ref).strip()] = round(spot, 4)
        return out

    def reference_results(self) -> list[dict]:
        """
        Resultados que el propio libro calculó (MTM, componente spot, etc.).

        Se usan como referencia de reconciliación: permiten comparar el motor
        contra la planilla operativa y detectar diferencias metodológicas.
        """
        ws = self._find_sheet(r'forwards\s+cordada')
        if ws is None:
            return []

        header_row, _ = self._locate_header(ws, r'^ref$')
        if header_row is None:
            return []

        headers = [str(c.value).strip().lower() if c.value else ''
                   for c in ws[header_row]]

        def col(pattern):
            rx = re.compile(pattern, re.IGNORECASE)
            for i, h in enumerate(headers):
                if rx.search(h):
                    return i
            return None

        idx = {
            'ref': col(r'^ref$'),
            'contraparte': col(r'contraparte'),
            'vcto': col(r'^vcto'),
            'monto': col(r'^monto$'),
            'spot_inicio': col(r'tipo de cambio al inicio'),
            'fwd_contrato': col(r'fwd contrato'),
            'dias': col(r'd[ií]as a vcto'),
            'fwd_bbg': col(r'fwd bbg'),
            'disc_rate': col(r'discount rate'),
            'disc_factor': col(r'factor de descuento'),
            'mtm': col(r'mtm'),
            'componente_spot': col(r'componente spot'),
        }
        if idx['ref'] is None or idx['mtm'] is None:
            return []

        out = []
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if idx['ref'] >= len(row) or row[idx['ref']] is None:
                continue
            rec = {}
            for key, i in idx.items():
                if i is None or i >= len(row):
                    rec[key] = None
                    continue
                v = row[i]
                rec[key] = _as_date(v) or (_as_float(v) if not isinstance(v, str) else v)
            if rec.get('mtm') is not None:
                out.append(rec)
        return out

    def close(self):
        self.wb.close()

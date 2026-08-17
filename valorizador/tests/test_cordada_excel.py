"""
Tests del lector del libro Cordada (`valorizador.services.cordada_excel`).

El libro se construye en memoria con openpyxl: no se depende de ningún archivo
externo. El cargador original hardcodeaba nombres de hoja y posiciones de celda
y fallaba en silencio si el libro del mes siguiente cambiaba de nombre; estos
tests fijan que la localización sea por patrón y que lo que no se encuentre
quede registrado como advertencia.
"""

import io
import unittest
from datetime import date, datetime

import openpyxl

from valorizador.services.cordada_excel import CordadaWorkbook

FWD_NODOS = [(1, 892.21), (2, 892.205), (8, 892.19), (62, 892.03)]
DESC_NODOS = [(92, 3.48231), (183, 3.61177), (365, 3.78017)]


def libro_cordada(
    *,
    nombre_valorizacion="Forwards Cordada 31",
    nombre_vigentes="FWD Vigentes al 31",
    con_curve_master=True,
    con_spot_inicio=True,
) -> io.BytesIO:
    """Arma un libro con la estructura del archivo Cordada real."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = nombre_valorizacion

    ws.append(["Fecha", datetime(2026, 5, 31)])
    ws.append([])
    encabezados = [
        "Ref", "Contraparte", "Vcto", "Monto",
        "Tipo de Cambio al Inicio" if con_spot_inicio else "Otra columna",
        "Fwd Contrato", "Días a Vcto", "Tipo de Cambio Fecha 31-05-2026",
        "Fwd BBG", "Discount Rate", "Factor de Descuento", "MTM", "Componente Spot",
    ]
    ws.append(encabezados)
    filas = [
        ("756929", "BTG Pactual", datetime(2026, 7, 7), 1_000_000, 887.71, 886.94,
         37, 892.89, 892.054193548387, 3.404064945054945, 0.9965655189778464,
         -5_096_628.947701437, -5_162_209.388305195),
        ("118039", "Bice", datetime(2026, 7, 13), 2_000_000, 894.25, 893.35,
         43, 892.89, 892.0483870967741, 3.412600769230769, 0.9959998685432817,
         2_592_812.5610144176, 2_709_119.6424377537),
    ]
    for f in filas:
        ws.append(list(f))

    if con_curve_master:
        cm = wb.create_sheet("CURVE MASTER")
        cm.append(["dia_FWDUSDCLP", "c_FWDUSDCLP", "dia_CLP423", "c_CLP423"])
        for i in range(max(len(FWD_NODOS), len(DESC_NODOS))):
            fila = []
            fila += list(FWD_NODOS[i]) if i < len(FWD_NODOS) else [None, None]
            fila += list(DESC_NODOS[i]) if i < len(DESC_NODOS) else [None, None]
            cm.append(fila)

    vig = wb.create_sheet(nombre_vigentes)
    vig.append(["Contraparte", "Folio", "Estado", "Operación", "Modalidad",
                "Fecha Inicio", "Fecha Vcto", "Monto", "Moneda", "Precio Fwd"])
    vig.append(["BTG Pactual", "756929", "Vigente", "Venta", "Compensacion",
                datetime(2026, 4, 7), datetime(2026, 7, 7), 1_000_000, "USD", 886.94])
    vig.append(["Bice", "118039", "Vigente", "Venta", "Compensacion",
                datetime(2026, 4, 13), datetime(2026, 7, 13), 2_000_000, "USD", 893.35])
    vig.append(["Bice", "999999", "Liquidado", "Compra", "Compensacion",
                datetime(2026, 1, 13), datetime(2026, 4, 13), 500_000, "USD", 880.00])

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream


class FechaYSpotTest(unittest.TestCase):
    """Localización de la fecha de valorización y del tipo de cambio."""

    def test_lee_la_fecha_y_el_spot_del_libro(self):
        libro = CordadaWorkbook(libro_cordada())
        fecha, spot = libro.valuation_date_and_spot()
        self.assertEqual(fecha, date(2026, 5, 31))
        self.assertAlmostEqual(spot, 892.89, places=4)
        libro.close()

    def test_encuentra_la_hoja_aunque_cambie_el_sufijo_del_nombre(self):
        """
        El nombre de la hoja lleva el día del mes: 'Forwards Cordada 30',
        'Forwards Cordada 31'... El original lo tenía hardcodeado.
        """
        libro = CordadaWorkbook(
            libro_cordada(nombre_valorizacion="Forwards Cordada 30-06")
        )
        fecha, spot = libro.valuation_date_and_spot()
        self.assertEqual(fecha, date(2026, 5, 31))
        self.assertIsNotNone(spot)
        libro.close()

    def test_avisa_cuando_no_encuentra_la_hoja_de_valorizacion(self):
        libro = CordadaWorkbook(libro_cordada(nombre_valorizacion="Resumen"))
        fecha, spot = libro.valuation_date_and_spot()
        self.assertIsNone(fecha)
        self.assertIsNone(spot)
        self.assertEqual(len(libro.warnings), 2)
        libro.close()


class CurvasTest(unittest.TestCase):
    """Lectura de los pares (dia_X, c_X) de la hoja CURVE MASTER."""

    def test_lee_todas_las_curvas_presentes(self):
        """El original leía sólo el primer par de columnas."""
        libro = CordadaWorkbook(libro_cordada())
        curvas = libro.curves()
        self.assertEqual(set(curvas), {"FWDUSDCLP", "CLP423"})
        self.assertEqual(
            curvas["FWDUSDCLP"],
            [{"tenor_days": d, "value": v} for d, v in FWD_NODOS],
        )
        self.assertEqual(
            curvas["CLP423"],
            [{"tenor_days": d, "value": v} for d, v in DESC_NODOS],
        )
        libro.close()

    def test_no_inventa_nodos_donde_una_curva_es_mas_corta(self):
        """CLP423 tiene tres nodos y FWDUSDCLP cuatro: no se rellena con ceros."""
        libro = CordadaWorkbook(libro_cordada())
        curvas = libro.curves()
        self.assertEqual(len(curvas["FWDUSDCLP"]), 4)
        self.assertEqual(len(curvas["CLP423"]), 3)
        libro.close()

    def test_avisa_si_falta_la_hoja_curve_master(self):
        libro = CordadaWorkbook(libro_cordada(con_curve_master=False))
        self.assertEqual(libro.curves(), {})
        self.assertTrue(any("CURVE MASTER" in w for w in libro.warnings))
        libro.close()


class ContratosTest(unittest.TestCase):
    """Lectura de la hoja de contratos vigentes."""

    def test_lee_los_contratos_vigentes_y_omite_los_liquidados(self):
        libro = CordadaWorkbook(libro_cordada())
        contratos = libro.contracts()
        self.assertEqual([c["folio"] for c in contratos], ["756929", "118039"])
        libro.close()

    def test_mapea_los_campos_del_contrato(self):
        libro = CordadaWorkbook(libro_cordada())
        c = next(x for x in libro.contracts() if x["folio"] == "118039")
        self.assertEqual(c["counterparty"], "Bice")
        self.assertEqual(c["side"], "Venta")
        self.assertEqual(c["notional"], 2_000_000.0)
        self.assertAlmostEqual(c["fwd_price"], 893.35, places=4)
        self.assertEqual(c["maturity_date"], date(2026, 7, 13))
        self.assertEqual(c["start_date"], date(2026, 4, 13))
        self.assertEqual(c["base_ccy"], "USD")
        self.assertEqual(c["quote_ccy"], "CLP")
        self.assertEqual(c["status"], "Vigente")
        libro.close()

    def test_toma_el_spot_al_inicio_de_la_hoja_de_valorizacion(self):
        """
        El cargador original asignaba a todos los contratos el spot **de hoy**
        como si fuera el spot del día en que se pactaron, con lo que el
        componente spot quedaba en cero por construcción.
        """
        libro = CordadaWorkbook(libro_cordada())
        por_folio = {c["folio"]: c for c in libro.contracts()}
        self.assertAlmostEqual(por_folio["756929"]["spot_inicio"], 887.71, places=4)
        self.assertAlmostEqual(por_folio["118039"]["spot_inicio"], 894.25, places=4)
        libro.close()

    def test_avisa_si_no_puede_mapear_el_spot_al_inicio(self):
        """Sin la columna, se importa igual pero con la advertencia explícita."""
        libro = CordadaWorkbook(libro_cordada(con_spot_inicio=False))
        contratos = libro.contracts()
        self.assertEqual(len(contratos), 2)
        self.assertTrue(all(c["spot_inicio"] == 0.0 for c in contratos))
        self.assertTrue(
            any("tipo de cambio al inicio" in w for w in libro.warnings),
            msg=f"avisos: {libro.warnings}",
        )
        libro.close()

    def test_avisa_si_falta_la_hoja_de_contratos(self):
        libro = CordadaWorkbook(libro_cordada(nombre_vigentes="Operaciones"))
        self.assertEqual(libro.contracts(), [])
        self.assertTrue(any("FWD Vigentes" in w for w in libro.warnings))
        libro.close()


class ResultadosDeReferenciaTest(unittest.TestCase):
    """Los resultados que la propia planilla calculó, para reconciliar."""

    def test_lee_los_resultados_por_folio(self):
        libro = CordadaWorkbook(libro_cordada())
        refs = {r["ref"]: r for r in libro.reference_results()}
        self.assertEqual(set(refs), {"756929", "118039"})
        self.assertAlmostEqual(refs["118039"]["mtm"], 2_592_812.5610144176, places=4)
        self.assertAlmostEqual(
            refs["118039"]["componente_spot"], 2_709_119.6424377537, places=4
        )
        self.assertAlmostEqual(refs["118039"]["fwd_bbg"], 892.0483870967741, places=8)
        self.assertAlmostEqual(refs["118039"]["disc_rate"], 3.412600769230769, places=8)
        self.assertEqual(refs["118039"]["dias"], 43)
        libro.close()

    def test_sin_hoja_de_valorizacion_no_hay_referencias(self):
        libro = CordadaWorkbook(libro_cordada(nombre_valorizacion="Resumen"))
        self.assertEqual(libro.reference_results(), [])
        libro.close()


if __name__ == "__main__":  # pragma: no cover
    unittest.main()
